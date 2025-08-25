#!/usr/bin/env python3
import os
import sys
import logging

# Configure logging for Railway
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

logger.info("=== Flask 앱 초기화 시작 ===")

# Load environment variables
try:
    from dotenv import load_dotenv
    load_dotenv()
    logger.info("✅ 환경변수 로드 완료")
except Exception as e:
    logger.warning(f"⚠️ dotenv 로드 실패: {e}")

# Import Flask components
try:
    from flask import Flask, render_template, request, jsonify, send_from_directory
    from werkzeug.utils import secure_filename
    import uuid
    from urllib.parse import urlparse
    from datetime import datetime
    logger.info("✅ Flask 모듈 import 완료")
except ImportError as e:
    logger.error(f"❌ Flask 모듈 import 실패: {e}")
    raise

# Import Supabase
try:
    from supabase import create_client, Client
    logger.info("✅ Supabase 모듈 import 완료")
except ImportError as e:
    logger.error(f"❌ Supabase 모듈 import 실패: {e}")
    raise

logger.info("Flask 앱 생성 중...")
app = Flask(__name__)

# 파일 업로드 설정 (Supabase Storage 전용)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
app.config['MAX_CONTENT_LENGTH'] = 1 * 1024 * 1024  # 1MB max file size

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Supabase 설정
logger.info("Supabase 연결 설정 중...")
supabase_url = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
supabase_key = os.getenv('NEXT_PUBLIC_SUPABASE_ANON_KEY')
supabase_service_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')

logger.info(f"Supabase URL 존재: {bool(supabase_url)}")
logger.info(f"Supabase Key 존재: {bool(supabase_key)}")

if not supabase_url or not supabase_key:
    logger.error("❌ Supabase 환경변수 누락")
    logger.error("필요한 환경변수: NEXT_PUBLIC_SUPABASE_URL, NEXT_PUBLIC_SUPABASE_ANON_KEY")
    supabase = None
    supabase_admin = None
else:
    try:
        # 일반 클라이언트 (데이터베이스용)
        supabase: Client = create_client(supabase_url, supabase_key)
        logger.info("✅ Supabase 클라이언트 생성 완료")
        
        # Storage용 관리자 클라이언트
        if supabase_service_key:
            supabase_admin: Client = create_client(supabase_url, supabase_service_key)
            logger.info("✅ Supabase 관리자 클라이언트 생성 완료")
        else:
            supabase_admin = supabase
            logger.warning("⚠️ Service Role Key 없음, anon key 사용")
        
        # 연결 테스트
        test_result = supabase.table('edutech_cards').select('id').eq('view', 1).limit(1).execute()
        logger.info(f"✅ Supabase 연결 테스트 성공 - 데이터: {bool(test_result.data)}")
    except Exception as e:
        logger.error(f"❌ Supabase 연결 실패: {e}")
        supabase = None
        supabase_admin = None

logger.info("=== Flask 앱 초기화 완료 ===")

@app.route('/')
def index():
    return render_template('view.html')

@app.route('/admin')
def admin():
    return render_template('admin.html')

# Health check endpoint for Railway
@app.route('/health')
def health_check():
    health_status = {
        'status': 'healthy',
        'supabase_connected': supabase is not None,
        'timestamp': os.environ.get('RAILWAY_DEPLOYMENT_ID', 'local'),
        'environment': 'railway' if os.environ.get('RAILWAY_ENVIRONMENT') else 'local',
        'python_version': sys.version.split()[0]
    }
    
    # Supabase 연결 상태 확인
    if supabase:
        try:
            test_result = supabase.table('edutech_cards').select('id').eq('view', 1).limit(1).execute()
            health_status['database_test'] = 'success'
            health_status['total_cards'] = len(test_result.data) if test_result.data else 0
        except Exception as e:
            health_status['database_test'] = f'failed: {str(e)}'
            logger.error(f"Health check database test failed: {e}")
    else:
        health_status['database_test'] = 'no_connection'
    
    return health_status

@app.route('/api/cards', methods=['GET', 'POST'])
def cards():
    if request.method == 'GET':
        try:
            print("=== API /api/cards GET 요청 받음 ===")
            
            if not supabase:
                print("ERROR: Supabase 연결 없음")
                return jsonify({'error': 'Database not configured'}), 500
                
            search = request.args.get('search', '')
            category = request.args.get('category', '')
            subject = request.args.get('subject', '')
            admin_view = request.args.get('admin', '')  # Admin can see all cards
            
            print(f"검색 조건 - search: '{search}', category: '{category}', subject: '{subject}', admin: '{admin_view}'")
            
            # Admin can see all cards, regular users only see approved ones
            if admin_view == 'true':
                query = supabase.table('edutech_cards').select('*')
            else:
                query = supabase.table('edutech_cards').select('*').eq('view', 1)
            
            if search:
                query = query.or_(f"webpage_name.ilike.%{search}%,user_summary.ilike.%{search}%,ai_summary.ilike.%{search}%")
            
            if category:
                query = query.eq('ai_category', category)
            
            if subject:
                query = query.contains('useful_subjects', [subject])
            
            # sort_order 필드가 있으면 우선 정렬, 없으면 created_at으로만 정렬
            try:
                query = query.order('sort_order', desc=False).order('created_at', desc=True)
            except Exception as sort_error:
                print(f"sort_order 필드 정렬 실패, created_at으로 대체: {sort_error}")
                query = query.order('created_at', desc=True)
            
            print("데이터베이스 쿼리 실행 중...")
            result = query.execute()
            
            card_count = len(result.data) if result.data else 0
            print(f"조회된 카드 수: {card_count}")
            
            if card_count > 0:
                print(f"첫 번째 카드: {result.data[0].get('webpage_name', 'Unknown')}")
            
            return jsonify(result.data or [])
            
        except Exception as e:
            print(f"ERROR: 카드 조회 실패 - {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': 'Failed to fetch cards'}), 500
    
    elif request.method == 'POST':
        try:
            data = request.json
            url = data.get('url')
            webpage_name = data.get('webpage_name')
            user_summary = data.get('user_summary', '')
            useful_subjects = data.get('useful_subjects', [])
            educational_meaning = data.get('educational_meaning', '')
            keyword = data.get('keyword', [])
            # JavaScript에서 이미 배열로 전송되므로 추가 처리 불필요
            if not isinstance(keyword, list):
                keyword = []
            
            if not url or not webpage_name:
                return jsonify({'error': 'URL and webpage_name are required'}), 400
            
            # 썸네일 URL 처리
            thumbnail_url = data.get('thumbnail_url', '')
            print(f"받은 썸네일 URL: '{thumbnail_url}'")  # 디버깅용
            if not thumbnail_url:
                thumbnail_url = f"https://via.placeholder.com/400x300?text={webpage_name}"
                print(f"기본 썸네일 URL 설정: '{thumbnail_url}'")  # 디버깅용
            
            new_card = {
                'url': url,
                'webpage_name': webpage_name,
                'user_summary': user_summary,
                'useful_subjects': useful_subjects,
                'educational_meaning': educational_meaning,
                'keyword': keyword,
                'thumbnail_url': thumbnail_url,
                'view': data.get('view', 1),  # 요청에서 view 값 가져오기, 기본값은 1 (admin에서 직접 추가시)
            }
            
            # sort_order 필드가 존재하는지 확인하고 설정
            try:
                # 새 카드의 sort_order는 0으로 설정 (맨 앞에 추가)
                new_card['sort_order'] = 0
                print(f"sort_order 설정: 0 (맨 앞에 배치)")
            except Exception as sort_error:
                print(f"sort_order 필드 설정 실패, 건너뜀: {sort_error}")
                # sort_order 필드가 없으면 그냥 추가하지 않음
            
            print(f"데이터베이스에 저장할 카드: {new_card}")  # 디버깅용
            result = supabase.table('edutech_cards').insert(new_card).execute()
            print(f"저장된 카드: {result.data[0] if result.data else 'None'}")  # 디버깅용
            return jsonify(result.data[0]), 201
            
        except Exception as e:
            print(f"Error creating card: {e}")
            if 'duplicate key' in str(e):
                return jsonify({'error': 'URL already exists'}), 409
            return jsonify({'error': 'Failed to create card'}), 500

@app.route('/api/cards/<int:card_id>', methods=['PUT', 'DELETE'])
def card_operations(card_id):
    if not supabase:
        return jsonify({'error': 'Database not configured'}), 500
        
    if request.method == 'PUT':
        try:
            data = request.json
            
            # 편집 비밀번호 확인
            password = data.get('password', '')
            EDIT_PASSWORD = "1"
            
            if password != EDIT_PASSWORD:
                return jsonify({'error': '편집 비밀번호가 일치하지 않습니다'}), 401
            
            url = data.get('url')
            webpage_name = data.get('webpage_name')
            user_summary = data.get('user_summary', '')
            useful_subjects = data.get('useful_subjects', [])
            educational_meaning = data.get('educational_meaning', '')
            keyword = data.get('keyword', [])
            if not isinstance(keyword, list):
                keyword = []
            
            if not url or not webpage_name:
                return jsonify({'error': 'URL and webpage_name are required'}), 400
            
            update_data = {
                'url': url,
                'webpage_name': webpage_name,
                'user_summary': user_summary,
                'useful_subjects': useful_subjects,
                'educational_meaning': educational_meaning,
                'keyword': keyword,
                'updated_at': 'now()'
            }
            
            # 썸네일 URL이 제공되면 업데이트
            thumbnail_url = data.get('thumbnail_url')
            if thumbnail_url is not None:
                update_data['thumbnail_url'] = thumbnail_url
            
            # view 필드 처리 (관리자 승인/거부)
            view_status = data.get('view')
            if view_status is not None:
                update_data['view'] = int(view_status)
            
            # 먼저 카드가 존재하는지 확인 (view 상태와 무관하게)
            check_result = supabase.table('edutech_cards').select('id').eq('id', card_id).execute()
            if not check_result.data:
                return jsonify({'error': 'Card not found'}), 404
            
            result = supabase.table('edutech_cards').update(update_data).eq('id', card_id).execute()
                
            return jsonify(result.data[0] if result.data else {'message': 'Updated successfully'})
            
        except Exception as e:
            print(f"Error updating card: {e}")
            return jsonify({'error': 'Failed to update card'}), 500
    
    elif request.method == 'DELETE':
        try:
            data = request.json or {}
            password = data.get('password', '')
            
            # 비밀번호 확인
            ADMIN_PASSWORD = "admin"
            
            if password != ADMIN_PASSWORD:
                return jsonify({'error': '비밀번호가 일치하지 않습니다'}), 401
            
            # 먼저 카드가 존재하는지 확인 (view 상태와 무관하게)
            check_result = supabase.table('edutech_cards').select('id, view').eq('id', card_id).execute()
            if not check_result.data:
                return jsonify({'error': 'Card not found'}), 404
            
            # 카드를 숨기기 (view=0으로 설정)
            result = supabase.table('edutech_cards').update({'view': 0}).eq('id', card_id).execute()
            
            return jsonify({'message': 'Card deleted successfully'})
            
        except Exception as e:
            print(f"Error deleting card: {e}")
            return jsonify({'error': 'Failed to delete card'}), 500


# OpenAI 분석 기능 제거됨

@app.route('/api/duplicate-check', methods=['POST'])
def duplicate_check():
    try:
        data = request.json
        url = data.get('url')
        
        if not url:
            return jsonify({'error': 'URL is required'}), 400
        
        domain = urlparse(url).hostname
        
        result = supabase.table('edutech_cards').select('id, webpage_name, url').eq('view', 1).ilike('url', f'%{domain}%').limit(5).execute()
        
        return jsonify({'duplicates': result.data or []})
        
    except Exception as e:
        print(f"Error checking duplicates: {e}")
        return jsonify({'error': 'Failed to check duplicates'}), 500

# 웹페이지 내용 추출 함수 제거됨 (OpenAI 분석 관련)

# Supabase Storage 설정
BUCKET_NAME = 'edutech-thumbnails'

@app.route('/api/upload-thumbnail', methods=['POST'])
def upload_thumbnail():
    try:
        if 'thumbnail' not in request.files:
            return jsonify({'error': 'No file selected'}), 400
        
        file = request.files['thumbnail']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if file and allowed_file(file.filename):
            # 파일 크기 체크 (1MB 제한)
            file.seek(0, 2)  # 파일 끝으로 이동
            file_size = file.tell()
            file.seek(0)  # 파일 시작으로 돌아가기
            
            if file_size > 1 * 1024 * 1024:  # 1MB
                return jsonify({'error': 'File size must be less than 1MB'}), 400
            
            # 고유한 파일명 생성
            file_extension = file.filename.rsplit('.', 1)[1].lower()
            unique_filename = f"{uuid.uuid4().hex}.{file_extension}"
            filename = secure_filename(unique_filename)
            
            # Supabase Storage에 업로드
            file_content = file.read()
            
            try:
                # Supabase Storage에 파일 업로드 (관리자 권한 사용)
                storage_client = supabase_admin if supabase_admin else supabase
                
                print(f"업로드 시도: {filename}, 크기: {len(file_content)} bytes")
                
                # 파일 업로드
                upload_result = storage_client.storage.from_(BUCKET_NAME).upload(
                    path=filename,
                    file=file_content,
                    file_options={"content-type": file.content_type}
                )
                
                print(f"업로드 결과 타입: {type(upload_result)}")
                print(f"업로드 결과: {upload_result}")
                
                # 업로드 성공 시 공개 URL 생성
                public_url = storage_client.storage.from_(BUCKET_NAME).get_public_url(filename)
                
                print(f"공개 URL: {public_url}")
                
                return jsonify({
                    'success': True,
                    'filename': filename,
                    'url': public_url
                })
                
            except Exception as storage_error:
                print(f"Supabase Storage 오류: {storage_error}")
                return jsonify({'error': f'Storage upload failed: {str(storage_error)}'}), 500
        else:
            return jsonify({'error': 'Invalid file type. Only PNG, JPG, JPEG, GIF, WebP allowed'}), 400
            
    except Exception as e:
        print(f"Error uploading thumbnail: {e}")
        return jsonify({'error': 'Failed to upload thumbnail'}), 500

# Excel 다운로드 엔드포인트
@app.route('/api/download-excel', methods=['POST'])
def download_excel():
    try:
        if not supabase:
            return jsonify({'error': 'Database not configured'}), 500
            
        data = request.json or {}
        password = data.get('password', '')
        
        # 비밀번호 확인 (관리자 비밀번호 사용)
        ADMIN_PASSWORD = "admin"
        
        if password != ADMIN_PASSWORD:
            return jsonify({'error': '비밀번호가 일치하지 않습니다'}), 401
        
        # 모든 visible 카드 데이터 가져오기
        result = supabase.table('edutech_cards').select('*').eq('view', 1).order('sort_order', desc=False).order('created_at', desc=True).execute()
        
        if not result.data:
            return jsonify({'error': '다운로드할 데이터가 없습니다'}), 404
        
        # Excel 파일 생성
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import Response
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "에듀테크 카드"
        
        # 헤더 설정
        headers = [
            'ID', '웹페이지 이름', 'URL', '간단 요약', '유용한 교과목', 
            '키워드', '교육적 의미', '생성일', '수정일', '정렬순서'
        ]
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 배열 필드를 안전하게 처리하는 함수
        def safe_join(value):
            try:
                print(f"safe_join 처리 중 - value: {value}, type: {type(value)}, repr: {repr(value)}")
                
                if value is None:
                    return ''
                    
                # 문자열인 경우 먼저 처리
                if isinstance(value, str):
                    return value.strip()
                
                # 리스트, 튜플 등 순회 가능한 객체인지 확인
                try:
                    # hasattr로 __iter__ 확인하고 string은 제외
                    if hasattr(value, '__iter__') and not isinstance(value, (str, bytes)):
                        # 실제로 순회해서 확인
                        filtered_items = []
                        for item in value:
                            if item is not None:
                                item_str = str(item).strip()
                                if item_str:  # 빈 문자열이 아닌 경우만 추가
                                    filtered_items.append(item_str)
                        return ', '.join(filtered_items)
                except TypeError:
                    # 순회 불가능한 객체인 경우
                    pass
                
                # 기타 타입은 문자열로 변환
                return str(value).strip() if value else ''
                
            except Exception as e:
                print(f"safe_join 오류: {e}, value: {value}, type: {type(value)}")
                return ''
        
        # 첫 번째 카드의 데이터 구조 디버깅
        if result.data:
            first_card = result.data[0]
            print(f"첫 번째 카드 디버깅:")
            print(f"  - useful_subjects: {first_card.get('useful_subjects')} (type: {type(first_card.get('useful_subjects'))})")
            print(f"  - keyword: {first_card.get('keyword')} (type: {type(first_card.get('keyword'))})")
            print(f"  - 전체 키: {first_card.keys()}")
        
        # 데이터 작성
        for row, card in enumerate(result.data, 2):
            try:
                ws.cell(row=row, column=1, value=card.get('id'))
                ws.cell(row=row, column=2, value=card.get('webpage_name', ''))
                ws.cell(row=row, column=3, value=card.get('url', ''))
                ws.cell(row=row, column=4, value=card.get('user_summary', ''))
                
                # 안전하게 배열 필드 처리 - 직접 구현
                useful_subjects = card.get('useful_subjects')
                if useful_subjects is None:
                    subjects_str = ''
                elif isinstance(useful_subjects, str):
                    subjects_str = useful_subjects
                elif hasattr(useful_subjects, '__len__') and len(useful_subjects) > 0:
                    try:
                        subjects_str = ', '.join([str(x) for x in useful_subjects if x])
                    except:
                        subjects_str = str(useful_subjects)
                else:
                    subjects_str = str(useful_subjects) if useful_subjects else ''
                ws.cell(row=row, column=5, value=subjects_str)
                
                keyword = card.get('keyword')
                if keyword is None:
                    keyword_str = ''
                elif isinstance(keyword, str):
                    keyword_str = keyword
                elif hasattr(keyword, '__len__') and len(keyword) > 0:
                    try:
                        keyword_str = ', '.join([str(x) for x in keyword if x])
                    except:
                        keyword_str = str(keyword)
                else:
                    keyword_str = str(keyword) if keyword else ''
                ws.cell(row=row, column=6, value=keyword_str)
                
                ws.cell(row=row, column=7, value=card.get('educational_meaning', ''))
                ws.cell(row=row, column=8, value=card.get('created_at', ''))
                ws.cell(row=row, column=9, value=card.get('updated_at', ''))
                ws.cell(row=row, column=10, value=card.get('sort_order', ''))
            except Exception as row_error:
                print(f"Row {row} 처리 오류: {row_error}")
                print(f"Card data: {card}")
                # 오류가 발생한 행은 건너뛰고 계속 진행
                continue
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 바이너리 데이터로 변환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = Response(
            output.getvalue(),
            headers={
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename=edutech_cards_{datetime.now().strftime("%Y%m%d")}.xlsx'
            }
        )
        return response
        
    except Exception as e:
        print(f"Excel 다운로드 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Excel 다운로드 중 오류가 발생했습니다'}), 500

# 카드 순서 업데이트 엔드포인트
@app.route('/api/cards/reorder', methods=['POST'])
def reorder_cards():
    try:
        if not supabase:
            return jsonify({'error': 'Database not configured'}), 500
            
        data = request.json or {}
        password = data.get('password', '')
        card_orders = data.get('card_orders', [])
        
        # 비밀번호 확인 (관리자 비밀번호 사용)
        ADMIN_PASSWORD = "admin"
        
        if password != ADMIN_PASSWORD:
            return jsonify({'error': '비밀번호가 일치하지 않습니다'}), 401
        
        if not card_orders:
            return jsonify({'error': '카드 순서 정보가 필요합니다'}), 400
        
        # 각 카드의 순서 업데이트
        for order_info in card_orders:
            card_id = order_info.get('id')
            sort_order = order_info.get('sort_order')
            
            if card_id and sort_order is not None:
                result = supabase.table('edutech_cards').update({
                    'sort_order': sort_order
                }).eq('id', card_id).execute()
        
        return jsonify({'message': '카드 순서가 성공적으로 업데이트되었습니다'})
        
    except Exception as e:
        print(f"카드 순서 업데이트 오류: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': '카드 순서 업데이트 중 오류가 발생했습니다'}), 500

# 로컬 파일 서빙 라우트 제거 (Supabase Storage 전용으로 변경)

# Railway 환경 확인
is_railway = os.environ.get('RAILWAY_ENVIRONMENT') is not None
deployment_id = os.environ.get('RAILWAY_DEPLOYMENT_ID', 'unknown')
logger.info(f"Railway 환경: {is_railway}")
if is_railway:
    logger.info(f"배포 ID: {deployment_id}")
    logger.info(f"Railway 서비스: {os.environ.get('RAILWAY_SERVICE_NAME', 'unknown')}")

# Startup summary
logger.info("=== 앱 초기화 완료 요약 ===")
logger.info(f"✅ Flask 앱: {app.name}")
logger.info(f"✅ Supabase 연결: {'성공' if supabase else '실패'}")
logger.info(f"✅ Supabase Storage 버킷: {BUCKET_NAME}")
logger.info(f"✅ 환경: {'Railway' if is_railway else '로컬'}")

if __name__ == '__main__':
    # 로컬 개발환경
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    logger.info(f"🚀 로컬 서버 시작: 포트 {port}, 디버그 {debug}")
    app.run(debug=debug, host='0.0.0.0', port=port)
else:
    # 프로덕션 환경 (Railway/gunicorn)
    logger.info("🚀 프로덕션 환경에서 실행 중")
    if is_railway:
        logger.info("Railway 플랫폼 배포 완료")